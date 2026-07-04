"""Excel export and import for adaptation measures."""

import io
from typing import BinaryIO

from geoalchemy2.shape import to_shape, from_shape
from openpyxl import Workbook, load_workbook
from shapely import wkt
from sqlalchemy.orm import Session

from app.data import catalog
from app.models.models import AdaptationMeasure, MeasureImpact


def export_measures_xlsx(db: Session, kommune_id: int) -> bytes:
    """Export all measures for a kommune as an Excel file.

    Returns xlsx bytes ready for download.
    """
    measures = (
        db.query(AdaptationMeasure)
        .filter(AdaptationMeasure.kommune_id == kommune_id)
        .all()
    )

    wb = Workbook()

    # ── Sheet 1: Measures ──
    ws = wb.active
    ws.title = "Maßnahmen"
    headers = [
        "ID", "Name", "Typ", "Umsetzungsjahr", "Beschreibung",
        "Konfiguration", "Geometrie (WKT)",
        "Investition (€)", "Wartung/Jahr (€)", "Nutzen/Jahr (€)",
        "Ø Risiko-Reduktion (Index-Pkt., Σ)",
        "Anzahl", "Einheit",
    ]
    ws.append(headers)

    for m in measures:
        geom_wkt = ""
        if m.geometry is not None:
            shape = to_shape(m.geometry)
            geom_wkt = shape.wkt

        # Aggregate impacts
        impacts = db.query(MeasureImpact).filter(MeasureImpact.measure_id == m.id).all()
        total_invest = sum(imp.costs.get("investment", 0) for imp in impacts)
        total_maint = sum(imp.costs.get("annual_maintenance", 0) for imp in impacts)
        total_savings = sum(
            sum(v for v in imp.savings.values()) for imp in impacts
        )
        total_didx = sum(
            sum(v for v in (imp.indicator_deltas or {}).values()) for imp in impacts
        )
        count = next((imp.costs["count"] for imp in impacts if "count" in imp.costs), None)
        count = count if count is not None else (m.config or {}).get("count")
        unit_label = catalog.MEASURES_BY_CODE.get(m.measure_type, {}).get("unit_label") or ""

        import json
        config_str = json.dumps(m.config or {}, ensure_ascii=False)

        ws.append([
            m.id,
            m.name,
            m.measure_type,
            m.implementation_year,
            m.description or "",
            config_str,
            geom_wkt,
            round(total_invest, 2),
            round(total_maint, 2),
            round(total_savings, 2),
            round(total_didx, 2),
            count if count is not None else "",
            unit_label,
        ])

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Zusammenfassung")
    ws2.append(["Kennzahl", "Wert"])
    ws2.append(["Anzahl Maßnahmen", len(measures)])
    ws2.append(["Gesamtinvestition (€)", sum(
        sum(imp.costs.get("investment", 0)
            for imp in db.query(MeasureImpact).filter(MeasureImpact.measure_id == m.id).all())
        for m in measures
    )])

    # Auto-width
    for ws_sheet in [ws, ws2]:
        for col in ws_sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_measures_xlsx(db: Session, kommune_id: int, file: BinaryIO) -> dict:
    """Import measures from an Excel file.

    Expected format: same as export (Sheet 'Maßnahmen' with columns matching export).
    Columns beyond the first 7 (Investition/Wartung/Nutzen/Risiko-Reduktion/Anzahl/
    Einheit) are optional and backward-compatible - older exports without the
    trailing "Anzahl"/"Einheit" columns still import fine.
    Returns summary of imported/skipped rows.
    """
    import json

    wb = load_workbook(file, read_only=True)
    ws = wb["Maßnahmen"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    imported = 0
    skipped = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        try:
            if len(row) < 7:
                errors.append(f"Zeile {i}: Zu wenige Spalten")
                skipped += 1
                continue

            _id, name, mtype, impl_year, desc, config_str, geom_wkt = row[:7]

            if not name or not mtype:
                errors.append(f"Zeile {i}: Name oder Typ fehlt")
                skipped += 1
                continue

            # Parse geometry
            if geom_wkt:
                shape = wkt.loads(geom_wkt)
                geom = from_shape(shape, srid=4326)
            else:
                errors.append(f"Zeile {i}: Keine Geometrie")
                skipped += 1
                continue

            # Parse config
            try:
                config = json.loads(config_str) if config_str else {}
            except (json.JSONDecodeError, TypeError):
                config = {}

            # Anzahl-Spalte (falls vorhanden) überschreibt einen evtl. in der
            # Konfigurations-JSON eingebetteten count-Wert bewusst - im Bulk-
            # Excel-Workflow ist die sichtbare Anzahl-Zelle die maßgebliche
            # Quelle. Das "friert" einen zuvor nur dynamisch berechneten
            # Richtwert-Default beim Reimport in einen expliziten Override ein
            # (kein Rückweg ohne manuelles Leeren der Zelle) - akzeptiert.
            if len(row) > 11 and row[11] not in (None, ""):
                try:
                    config["count"] = int(round(float(row[11])))
                except (TypeError, ValueError):
                    pass

            measure = AdaptationMeasure(
                kommune_id=kommune_id,
                name=str(name),
                measure_type=str(mtype),
                geometry=geom,
                config=config,
                implementation_year=int(impl_year) if impl_year else None,
                description=str(desc) if desc else None,
            )
            db.add(measure)
            imported += 1

        except Exception as e:
            errors.append(f"Zeile {i}: {str(e)[:200]}")
            skipped += 1

    db.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20],  # Limit error messages
    }


def export_parameters_xlsx(db: Session, kommune_id: int, kommune_name: str = "") -> bytes:
    """Export all model parameters for a kommune as Excel."""
    from datetime import datetime
    from app.services import parameter_registry

    params = parameter_registry.catalog_parameters()
    overrides = parameter_registry.load_db_overrides(db, kommune_id)
    merged = parameter_registry.merge_overrides(params, overrides)

    wb = Workbook()
    ws = wb.active
    ws.title = "Parameter"
    headers = [
        "ID", "Ebene", "Kategorie", "Bezeichnung", "Wert", "Default",
        "Einheit", "Quelle", "Override-Quelle", "Geändert",
    ]
    ws.append(headers)
    for p in merged:
        ws.append([
            p["id"],
            p["layer_code"],
            p["layer_category"],
            p["label"],
            p["value"],
            p["default_value"],
            p.get("unit", ""),
            p.get("source", ""),
            p.get("custom_source") or "",
            "ja" if p.get("overridden") else "nein",
        ])

    meta = wb.create_sheet("Metadaten")
    meta.append(["Kommune", kommune_name or str(kommune_id)])
    meta.append(["Exportdatum", datetime.utcnow().isoformat()])
    meta.append(["Modellversion", "KAP2 0.1.0"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
