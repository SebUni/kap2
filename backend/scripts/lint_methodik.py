#!/usr/bin/env python3
"""Deterministische Lints für Methodik-Berichte (Aufgabe §7).

Seit der ersten Review-Runde als Vorschlag im Ledger, jetzt gebaut — Anlass sind
die Ledger-Befunde **248 / 258 / 264**: dreimal wurden Statusspalten auf
„übernommen" gesetzt, während Berichtsstellen noch abgelöste Werte trugen. Diese
Fehlerklasse ist vollständig maschinell prüfbar und gehört deshalb nicht in eine
Selbstverpflichtung, sondern in die CI.

Geprüft wird (Aufgabe §7):

1. **Beispiel-Blöcke** ausführen — jedes Mini-Rechenbeispiel muss aufgehen.
2. **Zeichentabelle**: jede Zeile mit Wert *und* Herkunft; verbotene Formulierungen.
3. **Parameter-Blöcke**: alle neun Pflichtfelder; Kostensätze mit Preisstand.
4. **Preisstand-Einheitlichkeit** über alle Kostensätze eines Berichts.
5. **Quellen-Ratchet**: jede `source_ref` des Risikos mit URL, Archiv und Datum.
6. **Bericht ⇄ Registry**: jeder Parameter-Block-Wert muss der Registry-Spec
   entsprechen (Eiserne Regel 5 — Divergenz ist ein Ledger-Fall, kein stiller Fix).
7. **Revisionsrückstände**: Werte, die der Bericht selbst als abgelöst ausweist
   („bis Rev. N", „statt", Korrekturhistorie), dürfen außerhalb von Historie- und
   Log-Abschnitten nicht mehr als *geltende* Werte vorkommen.

Aufruf:
    python backend/scripts/lint_methodik.py 98          # ein Risiko
    python backend/scripts/lint_methodik.py             # alle Berichte
Rückgabe: 0 = grün, 1 = mindestens ein Lint rot.
"""
from __future__ import annotations

import glob
import os
import re
import sys

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT)
DOCS = os.path.abspath(os.path.join(ROOT, "..", "docs", "methodik"))

PFLICHTFELDER = ("id", "wert", "einheit", "band", "herkunft", "quelle",
                 "preisstand", "bandzuordnung", "endpunkt")
# Herkunft und Band je Parameter-Block (Befund 395): `registry_abgleich()` braucht
# sie, um Bloecke OHNE Registry-Spec einzuordnen — Pfad, abgeleitete Groesse
# oder ungekennzeichnet. Gefuellt von `parameter_bloecke()`.
BLOCK_META: dict[str, dict[str, object]] = {}
VERBOTEN = ("Platzhalter", "wird bei Implementierung", "wird später",
            "später hergeleitet", "wird nachgezogen")
# Abschnitte, in denen abgeloeste Werte legitim stehen (Historie, Log, Ledger-Bezug).
#
# REGEL (Befund 343): Jeder Eintrag benennt eine KLASSE von Zeilen, nie eine
# Einzelstelle. Der geloeschte Eintrag „statt Fällen" traf im gesamten Bericht
# genau eine Zeile — ausgerechnet die mit dem seit fuenf Runden beanstandeten
# abgeloesten Kopfgewichtungs-Wert. Eine Ausnahme, die genau einen Fund
# unterdrueckt, ist keine Ausnahme, sondern eine Whitelist fuer diesen Fund.
#
# EXPLIZITER HISTORIE-MARKER statt Heuristik (Befunde 343/353/375/383/392/402/405).
#
# Sechs Runden lang wurde geraten, ob eine Zeile einen abgeloesten Wert
# historisiert oder ihn behauptet: erst ueber Stichwoerter, dann ueber
# Umfeld-Fenster (70 -> 25 Zeichen), zuletzt ueber eine Adjazenz-Regel. Jede
# Verfeinerung schloss die zuletzt gefundene Luecke und oeffnete eine neue —
# die Adjazenz-Regel pruefte, was ZWISCHEN Vermerk und Wert steht, und liess
# einen NACHGESTELLTEN Vermerk bedingungslos gelten (Befund 405).
#
# Die Serie endet hier: Eine Zeile ist genau dann Historie, wenn sie es
# ausdruecklich SAGT. Der Marker wird bewusst gesetzt und laesst sich nicht
# versehentlich formulieren; wer einen abgeloesten Wert als geltend schreibt,
# muesste ihn wissentlich danebensetzen. Zwoelf Zeilen im gesamten Pruefgut
# tragen ihn — der Aufwand ist klein, die Zusicherung dafuer eindeutig.
# Gedeckte Fundstellen je Lauf. Sie werden am Ende von `pruefe_bericht()`
# AUSGEGEBEN (Befund 419: bis Runde 23 wurden sie gesammelt und nie gezeigt, sodass
# im Lauf nicht erkennbar war, welche Funde der Marker gedeckt hat).
UNTERDRUECKT: dict[str, list[tuple[str, int, str]]] = {}

HISTORIE_MARKER = "<!--hist-->"

# Der Marker ist keine bedingungslose Freigabe (Befund 419): Eine markierte Zeile
# muss zusaetzlich einen REVISIONSVERMERK tragen — »Rev. N«, »bis Rev. N«,
# »Rev.-N-Stand«, »Befund N« oder »Nr. N« (Entscheidungslog). Eine Zeile wie
# »Der Bundesschaden betraegt <abgeloester Wert> und gilt heute.« mit blossem
# Marker bleibt damit rot: Der Marker allein historisiert nichts, er bestaetigt
# nur, dass die Zeile ihren Revisionsvermerk absichtlich traegt.
REVISIONSVERMERK = re.compile(
    r"Rev\.\s?-?\s?\d|Rev\.-\d|bis Rev|Befunde? \d|Nr\. \d")

# Ratchet auf die Zahl der markierten Zeilen im gesamten Pruefgut (Befund 419):
# Wer eine Zeile neu markiert, traegt die Zahl hier bewusst nach — sonst ist der
# Marker eine stille Einzelfreigabe. Gezaehlt werden Zeilen mit Marker in allen
# gelesenen Quellen ausser der Lint-Datei selbst (ihre Konstantendefinition traegt
# den Marker per Definition; der Commit zu Runde 22 hatte sie faelschlich
# mitgezaehlt). Aktueller Bestand, Rev. 14 nach Runde 23 (gemessen): Bericht 21 —
# Korrekturhistorie §3.2 (7 Zeilen), §7-Historie-Kommentar (3), Entscheidungslog
# (11 Zeilen: Nr. 2, 16, 18, 19, 23–29); Anlage k_uv_herleitung.md (4: Schwellen-
# reihe ohne Schwelle, drei Verworfen-Zeilen); Code k_uv_herleitung.py (7: die
# Quellzeilen, die diese vier Anlagenzeilen erzeugen).
HISTORIE_MARKER_SOLL: dict[str, int] = {"98": 32}
MARKER_ZAEHLER: dict[str, int] = {}


def ist_historie(zeile: str) -> bool:
    """Traegt die Zeile den ausdruecklichen Historie-Marker?"""
    return HISTORIE_MARKER in zeile


def zaehle_marker(nr: str, src: str, quelle: str) -> None:
    """Zaehlt markierte Zeilen je Quelle fuer den Marker-Ratchet (Befund 419)."""
    if quelle == "Code lint_methodik.py":
        return
    MARKER_ZAEHLER[nr] = MARKER_ZAEHLER.get(nr, 0) + sum(
        1 for z in src.split("\n") if HISTORIE_MARKER in z)



class Lint:
    def __init__(self) -> None:
        self.fehler: list[str] = []
        self.ok: list[str] = []

    def pruefe(self, bedingung: bool, name: str, detail: str = "") -> None:
        if bedingung:
            self.ok.append(name)
        else:
            self.fehler.append(f"{name}: {detail}")


def beispiel_bloecke(src: str, lint: Lint) -> None:
    bloecke = re.findall(r"```python test: (\S+)\n(.*?)```", src, re.S)
    lint.pruefe(bool(bloecke), "Beispiel-Blöcke vorhanden", "keine gefunden")
    for name, code in bloecke:
        try:
            exec(compile(code, name, "exec"), {})  # noqa: S102
            # Ein Block ohne Zusicherung prueft nichts (Befund 298e).
            lint.pruefe("assert" in code, f"Beispiel-Block {name} hat assert",
                        "Block laeuft, prueft aber nichts")
            lint.ok.append(f"Beispiel-Block {name}")
        except Exception as exc:  # pragma: no cover
            lint.fehler.append(f"Beispiel-Block {name} rot: {exc}")


def zeichentabelle(src: str, lint: Lint) -> None:
    m = re.search(r"### 3\.5 Zeichentabelle.*?\n(.*?)\n### ", src, re.S)
    if not m:
        lint.fehler.append("Zeichentabelle: Abschnitt 3.5 nicht gefunden")
        return
    zeilen = [z for z in m.group(1).split("\n") if z.strip().startswith("|")][2:]
    lint.pruefe(bool(zeilen), "Zeichentabelle nicht leer")
    for z in zeilen:
        zellen = [c.strip() for c in z.strip().strip("|").split("|")]
        if len(zellen) != 4:
            lint.fehler.append(f"Zeichentabelle malformed: {z[:60]}")
            continue
        herkunft = zellen[3]
        # Befund 344(2): Ein blosses "[" akzeptierte auch "[offen]" als Herkunft.
        # Verlangt wird jetzt ein Register-/Herleitungs-Anker oder eine
        # nummerierte Quellenangabe [12] bzw. ein benannter Datensatz.
        hat = (any(k in herkunft for k in ("register:", "herleitung:", "Zensus",
                                           "berechnet", "Ergebnis", "Tabelle"))
               or bool(re.search(r"\[\d+\]", herkunft)))
        lint.pruefe(hat, f"Zeichentabelle {zellen[0][:24]}",
                    f"ohne Herkunft: {herkunft[:50]}")
    for wort in VERBOTEN:
        lint.pruefe(wort not in src, f"verbotene Formulierung „{wort}“",
                    "kommt im Bericht vor")


def parameter_bloecke(src: str, lint: Lint) -> tuple[dict[str, str], set[float]]:
    m = re.search(r"## 7 Parameter-Blöcke.*?\n(.*?)\n## 8 ", src, re.S)
    if not m:
        lint.fehler.append("Parameter-Blöcke: Kapitel 7 nicht gefunden")
        return {}, set()
    werte: dict[str, str] = {}
    preisstaende: set[str] = set()
    baender: set[float] = set()
    for blk in re.split(r"^parameter:$", m.group(1), flags=re.M)[1:]:
        pid = re.search(r"id:\s*(\S+)", blk)
        if not pid:
            lint.fehler.append("Parameter-Block ohne id")
            continue
        name = pid.group(1)
        fehlend = [f for f in PFLICHTFELDER
                   if not re.search(rf"^\s*{f}:", blk, re.M)]
        lint.pruefe(not fehlend, f"Parameter-Block {name}", f"fehlt {fehlend}")
        # Mehrzeilige dict-Werte ({mm: {...},\n c44: {...}}) vollstaendig einlesen
        # (Befund 384): Bis Rev. 14 endete der Wert am Zeilenende, sodass die
        # zweite Entitaet nie in den Abgleich kam.
        wert = re.search(r"^\s*wert:\s*(.+?)(?=\n\s*(?:einheit|band|herkunft):)",
                         blk, re.M | re.S)
        if wert:
            roh_wert = " ".join(z.split("#")[0].strip()
                                for z in wert.group(1).split("\n"))
            werte[name] = re.sub(r"\s+", " ", roh_wert).strip()
        bd = re.search(r"^\s*band:\s*\[([^\]]+)\]", blk, re.M)
        eigenes_band: list[float] = []
        if bd:
            for zahl in re.findall(r"[0-9]+\.[0-9]+", bd.group(1)):
                baender.add(float(zahl))
                eigenes_band.append(float(zahl))
        hk = re.search(r"^\s*herkunft:\s*(\S+)", blk, re.M)
        BLOCK_META[name] = {"herkunft": hk.group(1) if hk else "",
                            "band": eigenes_band}
        ps = re.search(r"^\s*preisstand:\s*\"?([^\s\"#]+)", blk, re.M)
        einheit = re.search(r"^\s*einheit:\s*\"?([^\s\"#]+)", blk, re.M)
        ist_kostensatz = bool(einheit and "EUR" in einheit.group(1).upper())
        if ps and ps.group(1) != "null":
            preisstaende.add(ps.group(1))
        # §3.3: Kostensaetze MUESSEN einen Preisstand tragen (Befund 298f).
        lint.pruefe(not ist_kostensatz or (ps and ps.group(1) != "null"),
                    f"Preisstand bei Kostensatz {name}",
                    "einheit nennt EUR, preisstand ist null")
    lint.pruefe(len(preisstaende) <= 1, "Preisstand einheitlich",
                f"mehrere Preisstände: {sorted(preisstaende)}")
    return werte, baender


def registry_abgleich(nr: str, werte: dict[str, str], lint: Lint) -> None:
    """Bericht ⇄ Registry (Eiserne Regel 5)."""
    try:
        from app.services.engine.impact.params import IMPACT_PARAM_SPECS
    except Exception as exc:  # pragma: no cover
        lint.fehler.append(f"Registry nicht ladbar: {exc}")
        return
    praefix = {"98": "uv."}.get(nr)
    if praefix is None:
        return
    specs = {s["key"]: s["value"] for s in IMPACT_PARAM_SPECS
             if str(s.get("risk", "")).startswith("EXPECTED_ANNUAL_UV")}
    uebersprungen: list[str] = []
    abgedeckt_dict: set[str] = set()
    for pid, roh in werte.items():
        if not pid.startswith(praefix):
            continue
        key = pid[len(praefix):]
        if key not in specs:
            uebersprungen.append(pid)     # zusammengesetzte Blöcke (dicts, Pfade)
            continue
        try:
            bericht = float(roh)
        except ValueError:
            continue
        registry = float(specs[key])
        lint.pruefe(abs(bericht - registry) < 1e-9,
                    f"Bericht ⇄ Registry {pid}",
                    f"Bericht {bericht} ≠ Registry {registry}")
    # Dict-Blöcke ({mm: …, c44: …}) auf die Einzel-Specs abbilden — sie machen
    # neun der vierzehn Blöcke aus und wurden bis Rev. 8 still übersprungen (275).
    for pid, roh in werte.items():
        if not roh.startswith("{"):
            continue
        basis = pid[len(praefix):] if pid.startswith(praefix) else pid
        for teil, zahl in re.findall(r"([a-z0-9_+]+):\s*([0-9.]+)", roh):
            for kandidat in (f"{basis}_{teil}", f"{basis}_{teil.replace('+', 'p')}"):
                if kandidat in specs:
                    lint.pruefe(abs(float(zahl) - float(specs[kandidat])) < 1e-9,
                                f"Bericht ⇄ Registry {pid}.{teil}",
                                f"Bericht {zahl} ≠ Registry {specs[kandidat]}")
                    break

    # Befund 344(4)/395: Uebersprungene Bloecke nicht still verschlucken — und
    # jeden Block ohne Spec EINORDNEN statt nur zaehlen. Ein Mutations-Sweep
    # (Runde 20) fand zwei Bloecke, die durch jeden Zweig fielen:
    #   * Pfad-Werte (`uv.ssd_delta_region`) → Existenz der Datei pruefen;
    #   * abgeleitete Groessen (`uv.r_out_sensitivitaet`, herkunft `herleitung:`)
    #     → der Wert muss im eigenen Band liegen (1,0 → 2,0 ist damit rot);
    #   * alles andere ohne Spec ist ungekennzeichnet → rot, namentlich.
    dict_bloecke, pfad_bloecke, abgeleitet, ungekennzeichnet = [], [], [], []
    for pid in uebersprungen:
        roh = werte[pid].strip().strip('"').strip("'")
        meta = BLOCK_META.get(pid, {})
        if roh.startswith("{"):
            dict_bloecke.append(pid)
            continue
        if pid == f"{praefix}voly":
            continue                 # Katalog-Parameter, eigener Abgleich unten (384b)
        if re.fullmatch(r"[\w./-]+\.(csv|npz|json|md|gpkg)", roh):
            pfad_bloecke.append(pid)
            voll = os.path.join(ROOT, "..", roh)
            lint.pruefe(os.path.exists(voll), f"Pfad-Block {pid} zeigt auf eine Datei",
                        f"{roh} existiert nicht")
            continue
        herkunft = str(meta.get("herkunft", ""))
        band = list(meta.get("band", []))
        try:
            zahl = float(roh)
        except ValueError:
            zahl = None
        if herkunft.startswith("herleitung:") and zahl is not None and len(band) >= 2:
            abgeleitet.append(pid)
            lint.pruefe(min(band) - 1e-9 <= zahl <= max(band) + 1e-9,
                        f"Abgeleiteter Block {pid} liegt im eigenen Band",
                        f"Wert {zahl} ausserhalb {band} — Herleitung {herkunft} "
                        f"traegt den Wert nicht")
            continue
        ungekennzeichnet.append(pid)
    lint.pruefe(not ungekennzeichnet, "Parameter-Blöcke ohne Registry-Spec gekennzeichnet",
                f"weder Spec noch Pfad noch `herleitung:`-Band: {ungekennzeichnet}")
    lint.ok.append(f"Bericht ⇄ Registry: {len(specs)} Specs; ohne Spec: "
                   f"{len(dict_bloecke)} Dict-Blöcke (unten abgeglichen) "
                   f"{dict_bloecke}, {len(pfad_bloecke)} Pfad-Blöcke {pfad_bloecke}, "
                   f"{len(abgeleitet)} abgeleitete {abgeleitet}")

    # Befund 384(a): VERSCHACHTELTE dict-Bloecke ({mm: {...}, c44: {...}}) wurden
    # vom flachen Parser nicht aufgeloest — `uv.i_raten_roh` fiel deshalb durch
    # JEDEN Abgleich, eine Mutation 24,7 -> 99,9 blieb gruen.
    for pid, roh in werte.items():
        if not roh.startswith("{") or "{" not in roh[1:]:
            continue
        basis = pid[len(praefix):] if pid.startswith(praefix) else pid
        for ent, inhalt in re.findall(r"([a-z0-9_]+):\s*\{([^}]*)\}", roh):
            for teil, zahl in re.findall(r"([a-z0-9_+\-]+):\s*([0-9.]+)", inhalt):
                t = teil.replace("+", "p").replace("-", "_")
                if t[0].isdigit():          # 20_64 -> a20_64, 85p -> a85p
                    t = "a" + t
                key = f"i_{ent}_{t}"
                if key in specs:
                    lint.pruefe(abs(float(zahl) - float(specs[key])) < 1e-9,
                                f"Bericht ⇄ Registry {pid}.{ent}.{teil}",
                                f"Bericht {zahl} ≠ Registry {specs[key]}")
                    abgedeckt_dict.add(key)

    # Befund 384(b): VOLY traegt zwei Drittel der Bundessumme, steht aber als
    # risikouebergreifender Parameter im Katalog, nicht in den UV-Specs — und
    # fiel damit durch den Risiko-Abgleich. Jetzt gegen den Katalogwert geprueft.
    voly_bericht = werte.get(f"{praefix}voly")
    if voly_bericht:
        try:
            from app.data.catalog import IMPACT_CATALOG  # noqa: F401
        except Exception:
            pass
        kat = os.path.join(ROOT, "app", "data", "catalog.py")
        if os.path.exists(kat):
            quelle = open(kat, encoding="utf-8").read()
            treffer = re.findall(r"160_?800(?:\.0)?", quelle)
            lint.pruefe(bool(treffer) and float(voly_bericht) == 160800.0,
                        "Bericht ⇄ Katalog uv.voly",
                        f"Bericht {voly_bericht}, Katalog fuehrt "
                        f"{'160800' if treffer else 'keinen passenden Wert'}")

    # Befund 364: VOLLSTAENDIGKEIT — jede Registry-Spec braucht einen
    # Parameter-Block. Bis Rev. 14 deckten die Bloecke nur 15 von 28 Specs ab; die
    # uebrigen 13 (Inzidenzraten, or_out, qbar_out, r_out_enabled) wurden nie
    # gegen den Bericht abgeglichen, sodass eine Mutation dort unbemerkt blieb.
    abgedeckt = set()
    for pid, roh in werte.items():
        basis = pid[len(praefix):] if pid.startswith(praefix) else pid
        abgedeckt.add(basis)
        if roh.startswith("{"):
            for teil, _ in re.findall(r"([a-z0-9_+\-]+):\s*([0-9.]+)", roh):
                abgedeckt.add(f"{basis}_{teil.replace('+', 'p').replace('-', '_')}")
    abgedeckt |= abgedeckt_dict
    ohne_block = sorted(k for k in specs if k not in abgedeckt)
    lint.pruefe(not ohne_block, "Parameter-Block je Registry-Spec",
                f"{len(ohne_block)} Specs ohne Block im Bericht: {ohne_block[:8]}")

    # Befund 344(1): Jeder SYMBOLE-Eintrag muss einem Registry-Key entsprechen.
    # Der Eintrag `voly` war wirkungslos, weil die UV-Specs keinen solchen Key
    # fuehren — `soll is None ⇒ continue` liess ihn stumm durchfallen, und der
    # Umsetzungsnachweis behauptete eine Wirkung, die es nicht gab.
    for sym in SYMBOLE.get(nr, {}):
        lint.pruefe(sym in specs, f"SYMBOLE-Eintrag {sym} trifft eine Registry-Spec",
                    f"kein Key `{sym}` in den Specs — Eintrag ist wirkungslos")
    lint.ok.append(f"SYMBOLE-Abdeckung: {len(SYMBOLE.get(nr, {}))} von "
                   f"{len(specs)} Specs")


# Berichtsbezeichnung je Registry-Parameter — Grundlage des Rückstands-Checks.
# Nur Parameter, die im Fließtext mit ihrem Zahlenwert genannt werden.
# Zahlen, die in einer Formelzeichen-Zeile legitim NEBEN dem Wert stehen dürfen
# (Zwischenergebnisse der dokumentierten Herleitung). Alles andere in derselben
# Größenordnung ist ein Rückstandsverdacht — bewusst streng.
# Zahlen, die in einer Formelzeichen-Zeile legitim NEBEN dem Wert stehen dürfen.
# ACHTUNG (Befund 298): Hier gehören **nur echte Zwischenergebnisse der aktuellen
# Herleitung** hinein — niemals abgelöste Vorgängerwerte. Eine Whitelist, die die
# Revisionshistorie enthält, entwertet die Negativprüfung vollständig; genau das war
# bis Rev. 10 der Fall. Jeder Eintrag braucht einen Kommentar, WOHER er stammt.
# Werte, die im Lauf dieses Berichts ABGELÖST wurden. Sie dürfen im geltenden Teil
# (alles außer Kopfvermerk, Korrekturhistorie und Entscheidungslog) nicht mehr
# vorkommen — weder in Prosa noch in Golden-Test-Kommentaren noch in Anlagen.
#
# Die Liste steht bewusst IM LINT und nicht im Bericht (Befund 286): Sie darf nicht
# von der Pflege der Korrekturhistorie abhaengen. Wer einen Modellwert aendert, traegt
# den alten hier ein — ein Handgriff, den der Review pruefen kann.
# HINWEIS zur Pflege: 5,65 %/Dek. stand bis Rev. 13 hier — es ist seit Rev. 14
# wieder ein GELTENDER Wert: §3.2 benennt ihn als den Globalstrahlungstrend, den
# die Abstract-Formulierung „about twice" implizieren wuerde, und stellt ihn dem
# publizierten Tab.-4-Wert 4,6 gegenueber. §3.8 verlangt genau diese Benennung.
# Belegt durch: grep -q "5,65 %/Dek" docs/methodik/98_uv_schaedigungen.md
#
# HINWEIS zur Pflege: 0,6320 stand bis Rev. 13 hier — es ist seit Rev. 14 wieder
# ein GELTENDER Wert: die Anlage k_uv_herleitung.md misst und druckt ihn als
# Trendgewichtungs-Sensitivitaet. Ein Wert gehoert nur in diese Liste, solange er
# nirgends mehr gilt; sonst meldet der Lint dauerhaft einen Rueckstand, den es
# nicht gibt. Belegt durch: grep -q "0.6320" backend/data/kalibrierung/k_uv_herleitung.md
ABGELOESTE_WERTE: dict[str, tuple[str, ...]] = {
    "98": (
        "0,8434", "0,7562", "0,5782", "0,6667", "0,6735", "0,6736", "0,7216",
        "0,6323", "0,6774", "0,6843", "0,7289", "0,6828", "0,6854",
        "/5,81", "÷ 5,81", "4,9/5,81", "6,48 %/Dek", "4,32 %/Dek",
        "0,3656", "0,9187", "0,3427", "1,0044", "0,3671", "1,0760", "0,4336",
        "0,3709", "1,0870",
        "3,69 %", "4,25 %", "4,30 %", "4,61 %", "4,65 %", "4,83 %", "4,95 %",
        "5,38 %",
        "275 Mio", "317 Mio", "320 Mio", "343 Mio", "360 Mio", "367 Mio",
        "401 Mio", "YLL 1.141", "YLL 1.315", "YLL 1.329", "YLL 1.423", "YLL 1.438",
        "YLL 1.492", "YLL 1.521", "YLL 1.664", "YLL ≈ 1.141", "YLL ≈ 1.315",
        "YLL ≈ 1.329", "YLL ≈ 1.423", "YLL ≈ 1.438", "YLL ≈ 1.492",
        "10.808", "347 Mio", "5,01 %",
        # Abgeleitete Plausibilisierungswerte (Befund 418): 8,51 % x 0,7289 = 6,2 %
        # bzw. 2,1 %/Dek. gehoeren zum abgeloesten k_UV = 0,7289 (Rev. 9); geltend
        # sind 6,06 % und 2,02 %/Dek. Bewusst mit Praefix, weil »6,2 %« an
        # anderer Stelle legitim vorkommt (Behandlungs-€-Anteil an der KKR).
        "≈ 6,2 %", "~ 6,2 %", "2,1 %/Dek"),
}

ZWISCHENWERTE: dict[str, dict[str, tuple[float, ...]]] = {
    "98": {
        # Nur die Glieder der GELTENDEN k_UV-Kette:
        # 1,0652 = Stationsquotient 4,9/4,6 ([31] Tab. 2 / Tab. 4)
        # 0,6683 = Rasterquotient (Fallgewichtung, €-gewichtetes Mittel)
        # 0,6674 / 0,6689 = derselbe Quotient je Entität (MM / C44)
        # 0,6811 = Rasterquotient an der Messzelle Bochum
        "k_uv": (1.0652, 0.6683, 0.6674, 0.6689, 0.6811),
        # Stützstellen der Sterbetafel 2022/2024
        "l_rest_mm": (10.9187, 10.3350, 9.7311),
        "l_rest_c44": (5.0374, 5.9397, 5.4745),
    },
}

# Formelzeichen je Registry-Key. Jeder Eintrag MUSS einem Key der Risiko-Specs
# entsprechen — `registry_abgleich` prueft das (Befund 344). VOLY steht hier
# bewusst nicht: Es ist ein risikouebergreifender Parameter der #95-Kette und
# gehoert dort geprueft, nicht in den UV-Specs.
SYMBOLE: dict[str, tuple[str, ...]] = {
    "98": {
        "k_uv": (r"k_\{\\text\{UV\}\}", r"\bk_UV\b"),
        "a_attr": (r"a_\{\\text\{attr",),
        "w_scc": (r"w_\{\\text\{SCC\}\}",),
        "l_rest_mm": (r"\\bar L_\{\\text\{MM\}\}",),
        "l_rest_c44": (r"\\bar L_\{\\text\{C44\}\}",),
        "qbar_out": (r"\\bar q_\{\\text\{out\}\}",),
        "or_out": (r"\bOR\b",),
        "s_komforttag": (r"\bs\b(?=.*Komforttag)",),
        "c_kal_mm": (r"c_\{\\text\{kal,MM\}\}",),
        "c_kal_c44": (r"c_\{\\text\{kal,C44\}\}",),
    },
}


def abgeloeste_werte(nr: str, src: str, lint: Lint, quelle: str = "Bericht") -> None:
    """Abgeloeste Werte im geltenden Teil (Befunde 283/294/295/296).

    Der zielgenaue Check fuer die Fehlerklasse, die diesen Lauf dominiert hat: Nach
    einer Wertaenderung bleiben einzelne Fundstellen stehen — in Prosa, in
    Golden-Test-Kommentaren, in Anlagen. Die Suche laeuft ueber die im Lint gepflegte
    Liste, ist also unabhaengig von der Korrekturhistorie.
    """
    werte = ABGELOESTE_WERTE.get(nr)
    if not werte:
        lint.ok.append(f"Abgeloeste Werte ({quelle}: keine Liste hinterlegt)")
        return
    # Kapitel 8 NICHT abschneiden (Befund 313): Auch der Quellenblock traegt
    # wertetragende Zahlen und hatte Rueckstaende.
    aktuell = src
    # Selbstpruefung (Befund 347): Der Lint liest auch sich selbst. Die Definition
    # von ABGELOESTE_WERTE fuehrt die verbotenen Werte per Definition — nur dieser
    # eine Block ist ausgenommen, nicht die Datei. Kommentare, ZWISCHENWERTE und
    # Code bleiben im Pruefbereich; genau dort standen Rueckstaende.
    if quelle == "Code lint_methodik.py":
        aktuell = re.sub(r"ABGELOESTE_WERTE.*?\n\}\n",
                         lambda m: "\n" * m.group(0).count("\n"), aktuell, flags=re.S)
    # Vorkommen NUR im geprueften Inhalt zaehlen, nicht in der Lint-Datei selbst
    # (Befund 353): Dort steht jede Ausnahme mit Name UND Muster, was ein
    # Grundrauschen von 2 Treffern je Ausnahme erzeugt haette — knapp unter
    # MINDESTVORKOMMEN, aber genug, um eine echte Einzelfreigabe ueber die
    # Schwelle zu heben und den Ratchet damit wirkungslos zu machen.
    treffer = 0
    zaehle_marker(nr, aktuell, quelle)
    # KEINE Stichwort- und KEINE Abschnitts-Heuristik mehr (Befund 414): Bis Runde 23
    # galten Zeilen mit »Korrekturhistorie«, »Historie:« oder »abgelöst durch«
    # irgendwo im Text sowie ganze Abschnitte (»Verworfene …«, »Schwelle …«) als
    # Historie — »Ungeachtet der Korrekturhistorie gilt k_UV = <abgeloester
    # Wert>.« lief deshalb gruen. Historie ist jetzt ausschliesslich, was den Marker traegt;
    # Entscheidungslog-Zeilen, Korrekturhistorie und Verworfen-Listen sind
    # zeilenweise markiert.
    # Blockquote-Ausnahme NUR fuer den Kopfvermerk (Befund 345): Der mehrzeilige
    # Revisionsstand vor der ersten Kapitelueberschrift nennt abgeloeste Werte zu
    # Recht. Ab der ersten „## "-Ueberschrift sind Blockquotes normaler Berichtstext
    # — insbesondere die INFOKAESTEN in §6, die der Nutzer im Produkt sieht und die
    # bis Rev. 13 als einzige Berichtsteile ohne jede Wertkontrolle blieben.
    kopf_ende = aktuell.find("\n## ")
    if kopf_ende < 0:
        kopf_ende = len(aktuell)
    offset = 0
    for i, zeile in enumerate(aktuell.split("\n"), start=1):
        zeilen_start, offset = offset, offset + len(zeile) + 1
        if zeile.lstrip().startswith(">") and zeilen_start < kopf_ende:
            continue
        for wert in werte:
            # Anlagen schreiben Dezimalzahlen mit PUNKT (f-Strings), der Bericht mit
            # Komma — beide Formen pruefen (Befund 312).
            formen = {wert}
            if re.fullmatch(r"[0-9]+,[0-9]+", wert):
                formen.add(wert.replace(",", "."))
            # Wortgrenzen: ein vierstelliger Nachkommawert darf nicht in einer
            # laengeren Zahl mit demselben Ende anschlagen.
            treffer_pos = None
            for f in formen:
                m = re.search(r"(?<![0-9.,])" + re.escape(f), zeile)
                if m:
                    treffer_pos = m.start()
                    break
            if treffer_pos is None:
                continue
            # Historie-Ausnahme NUR im Umfeld der Fundstelle pruefen (Befund 312):
            # Lange Anlagen-Zeilen enthalten weiter hinten oft ein "ergaebe sich",
            # das sonst die ganze Zeile — und damit einen echten Rueckstand —
            # entschuldigt.
            # Befund 383: In BLOCKQUOTES ab Kapitel 1 gilt KEINE Historie-Ausnahme.
            # Das sind die Infokaesten des Produkts (§3.6: Berichtstext), und die
            # verbliebene Ausnahme „Rev. N:" ist breit genug, dass dort ein
            # beliebiger Revisionsverweis im Umfeld jeden abgeloesten Wert deckte.
            # Der Kopfvermerk bleibt ueber `zeilen_start < kopf_ende` ausgenommen.
            if zeile.lstrip().startswith(">"):
                lint.fehler.append(
                    f"Abgeloester Wert im {quelle}, Zeile {i} (Infokasten): "
                    f"{wert} — {zeile.strip()[:60]}")
                treffer += 1
                continue
            # Historie NUR ueber den expliziten Marker (Befund 405). Die
            # frueheren Heuristiken — Stichwortliste, Umfeld-Fenster, zuletzt
            # eine Adjazenz-Regel — haben ueber sechs Runden je die zuletzt
            # gefundene Luecke geschlossen und eine neue geoeffnet.

            if ist_historie(zeile):
                if not REVISIONSVERMERK.search(zeile):
                    # Befund 419: Marker ohne Revisionsvermerk ist keine Historie.
                    lint.fehler.append(
                        f"Historie-Marker ohne Revisionsvermerk im {quelle}, "
                        f"Zeile {i}: {wert} — {zeile.strip()[:60]}")
                    treffer += 1
                    continue
                UNTERDRUECKT.setdefault(HISTORIE_MARKER, []).append(
                    (quelle, i, wert))
                continue
            lint.fehler.append(
                f"Abgeloester Wert im {quelle}, Zeile {i}: {wert} — "
                f"{zeile.strip()[:65]}")
            treffer += 1
    if not treffer:
        lint.ok.append(f"Abgeloeste Werte ({quelle}: {len(werte)} geprueft)")


def revisionsrueckstaende(nr: str, src: str, baender: set[float],
                          lint: Lint) -> None:
    """Abgelöste Werte, die als *geltend* im Bericht stehen (Befunde 248/258/264).

    Zwei Prüfungen, die sich gegenseitig absichern:

    * **Positiv** (registry-basiert, nicht zirkulär): Der aktuelle Registry-Wert muss
      im geltenden Teil des Berichts mindestens einmal in einer Zeile mit dem
      zugehörigen Formelzeichen vorkommen. Fällt weg, sobald der Bericht bei einer
      Wertänderung nicht nachgezogen wurde.
    * **Negativ**: Werte, die die Korrekturhistorie ausdrücklich als abgelöst
      ausweist, dürfen außerhalb von Historie-, Log- und Kopfzeilen nicht mehr in
      einer Zeile mit demselben Formelzeichen stehen.

    Bandgrenzen und Zwischenwerte (Quotienten, Stützstellen) lösen nichts aus — sie
    sind legitime Zahlen in denselben Zeilen.
    """
    symbole = SYMBOLE.get(nr)
    if not symbole:
        lint.ok.append("Revisionsrückstände (kein Symbol-Mapping für dieses Risiko)")
        return
    try:
        from app.services.engine.impact.params import IMPACT_PARAM_SPECS
    except Exception as exc:  # pragma: no cover
        lint.fehler.append(f"Rückstands-Check: Registry nicht ladbar: {exc}")
        return
    specs = {s["key"]: s["value"] for s in IMPACT_PARAM_SPECS
             if str(s.get("risk", "")).startswith("EXPECTED_ANNUAL_UV")}

    # KEINE Negativmenge aus der Korrekturhistorie mehr (Befund 286): Sie hängt
    # davon ab, dass der Autor die Historie pflegt — genau die Disziplin, die der
    # Lint ersetzen soll. Stattdessen gilt: JEDE Zahl in einer Zeile mit genau
    # einem Formelzeichen muss der Registry-Wert, eine Bandgrenze oder ein
    # ausdrücklich erlaubter Zwischenwert sein.
    # GELTUNGSBEREICH (Befund 344/5): Diese Pruefung endet bewusst vor Kapitel 8,
    # `abgeloeste_werte()` schliesst es bewusst ein — die beiden Bereiche sind
    # verschieden, weil die Pruefungen Verschiedenes messen:
    #   * hier wird ein Wert NUR beanstandet, wenn er in einer Zeile mit seinem
    #     Formelzeichen steht. Im Quellenverzeichnis stehen Zahlen ohne
    #     Formelzeichen-Kontext (Seitenzahlen, Jahrgaenge, DOI-Fragmente), was
    #     reine Fehlalarme ergaebe.
    #   * `abgeloeste_werte()` sucht die Werte selbst, ohne Zeichenbindung, und
    #     muss Kapitel 8 deshalb einbeziehen — dort standen reale Rueckstaende
    #     (Befund 313).
    aktuell = src.split("## 8 Quellen")[0]
    # LaTeX schreibt Dezimalzahlen als 0{,}6736 — ohne Normalisierung sieht der
    # Lint genau die Definitionsgleichungen NICHT (Befund 274).
    zeilen = [z.replace("{,}", ",").replace("{.}", ".")
              for z in aktuell.split("\n")]
    # Kopfvermerk = Blockquote VOR der ersten Kapitelueberschrift (Befund 345):
    # dort stehen Revisionsstand und abgeloeste Werte zu Recht. Ab Kapitel 1 ist
    # ein Blockquote normaler Berichtstext (Infokaesten) — keine Ausnahme mehr
    # (Befund 414: die pauschale Blockquote-Ausnahme dieser Funktion hatte
    # aufgehoben, was Befund 383 fuer die Infokaesten gerade abgeschafft hatte).
    kopf_ende_zeile = next((i for i, z in enumerate(zeilen, start=1)
                            if z.startswith("## ")), len(zeilen) + 1)

    def im_kopf(i: int, zeile: str) -> bool:
        return zeile.lstrip().startswith(">") and i < kopf_ende_zeile

    for key, muster in symbole.items():
        soll = specs.get(key)
        if soll is None or not isinstance(soll, (int, float)):
            continue
        gefunden = False
        for i, zeile in enumerate(zeilen, start=1):
            if not any(re.search(m, zeile) for m in muster):
                continue
            # Befund 298: `#`-Zeilen sind die Kommentare der Golden-Test-Blöcke —
            # sie waren pauschal ausgenommen und trugen deshalb Rückstände.
            # Historie NUR ueber den expliziten Marker (Befunde 405/414). Bis
            # Runde 23 galten hier noch »Rev. N« am Zeilenende, eine Stichwort-
            # liste und jeder Blockquote — »k_UV = 0,9000 (Rev. 8)« lief gruen.
            historie = ist_historie(zeile) or im_kopf(i, zeile)
            # Positivprüfung braucht auch kurze Werte (0,75 · 1,45), die
            # Negativprüfung bleibt bei >= 3 Nachkommastellen (weniger Rauschen).
            zahlen = {float(r.replace(",", "."))
                      for r in re.findall(r"[0-9]+[,.][0-9]{1,6}", zeile)}
            lange = {float(r.replace(",", "."))
                     for r in re.findall(r"[0-9]+[,.][0-9]{3,6}", zeile)}
            if any(abs(z - soll) < 1e-6 for z in zahlen) and not historie:
                gefunden = True
            if historie:
                continue
            # Sammel-Zeilen (mehrere Formelzeichen) lassen sich nicht zuordnen.
            if sum(1 for ms in symbole.values()
                   if any(re.search(m, zeile) for m in ms)) > 1:
                continue
            for z in lange:
                if abs(z - soll) < 1e-6:
                    continue
                if any(abs(z - b) < 1e-9 for b in baender):
                    continue                       # Bandgrenze
                if any(abs(z - w) < 1e-9 for w in ZWISCHENWERTE.get(nr, {}).get(key, ())):
                    continue                       # dokumentierter Zwischenwert
                if 0.5 <= z / soll <= 2.0:
                    lint.fehler.append(
                        f"Rückstandsverdacht Zeile {i}: {key} — Zahl {z} in der "
                        f"Größenordnung des Werts {soll}, aber weder Registry-Wert "
                        f"noch Bandgrenze noch erlaubter Zwischenwert; "
                        f"{zeile.strip()[:60]}")
        lint.pruefe(gefunden, f"Registry-Wert im Bericht: {key} = {soll}",
                    "kommt in keiner Zeile mit dem Formelzeichen vor — "
                    "Bericht nach einer Wertänderung nicht nachgezogen?")
        # Definitionsgleichungen: eine hervorgehobene Zahl (**x** / \mathbf{x})
        # in einer Zeile mit Symbol UND "=" MUSS der Registry-Wert sein. Das
        # fängt den Fall, den die Positivprüfung durchlässt: Der richtige Wert
        # steht irgendwo, die Definitionsgleichung trägt aber den alten (274).
        for i, zeile in enumerate(zeilen, start=1):
            if ist_historie(zeile) or im_kopf(i, zeile):
                continue
            if "=" not in zeile or not any(re.search(m, zeile) for m in muster):
                continue
            # Zeilen, die mehrere Formelzeichen zusammenfassen (Sammel-Zeilen der
            # Zeichentabelle), lassen sich keinem Wert eindeutig zuordnen.
            if sum(1 for ms in symbole.values()
                   if any(re.search(m, zeile) for m in ms)) > 1:
                continue
            for roh in re.findall(r"(?:\*\*|\\mathbf\{)([0-9]+[,.][0-9]{1,6})",
                                  zeile):
                wert = float(roh.replace(",", "."))
                if any(abs(wert - b) < 1e-9 for b in baender):
                    continue                     # Bandgrenzen sind legitim
                if abs(wert - soll) / soll < 0.01:
                    continue                     # gerundete Anzeige desselben Werts
                if abs(wert - soll) > 1e-6 and 0.5 <= wert / soll <= 2.0:
                    lint.fehler.append(
                        f"Definitionsgleichung Zeile {i}: {key} = **{roh}**, "
                        f"Registry sagt {soll} — {zeile.strip()[:60]}")


def quellen_ratchet(lint: Lint) -> None:
    try:
        from app.data.sources import SOURCE_REFERENCES
        from app.services.engine.impact.params import IMPACT_PARAM_SPECS
    except Exception as exc:  # pragma: no cover
        lint.fehler.append(f"Quellen-Ratchet nicht prüfbar: {exc}")
        return
    refs = {r for s in IMPACT_PARAM_SPECS for r in (s.get("source_refs") or [])}
    for ref in sorted(refs):
        eintrag = SOURCE_REFERENCES.get(ref)
        if not eintrag:
            lint.fehler.append(f"Quelle {ref} fehlt im Register")
            continue
        fehlend = [k for k in ("url", "archive_url", "accessed")
                   if not eintrag.get(k)]
        lint.pruefe(not fehlend, f"Quelle {ref}", f"ohne {fehlend}")


# Knoten je Risiko in der Schadensbaum-Arbeitsmappe (§7: „Skript liest die xlsx").
XLSX = os.path.abspath(os.path.join(ROOT, "..", "docs", "Schadensbaum",
                                    "KWRA-Schadensbaum_X_UBA-klimawirkungsketten.xlsx"))
RISIKO_KNOTEN = {"98": "W186"}


def knoten_abgleich(nr: str, src: str, lint: Lint) -> None:
    """Knoten der Arbeitsmappe ⇄ Knoten-Bilanz des Berichts (Aufgabe §7, LF 1/14).

    Bis Rev. 9 lief dieser Abgleich nur im Review von Hand (Befund 287). Er liest
    die xlsx und prüft, dass jeder Input-Knoten des Risikos in der Knoten-Bilanz
    vorkommt und kein Knoten behauptet wird, den die Mappe nicht führt.
    """
    knoten_id = RISIKO_KNOTEN.get(nr)
    if not knoten_id:
        lint.ok.append("Knoten-Abgleich (kein Mapping für dieses Risiko)")
        return
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        lint.fehler.append("Knoten-Abgleich: openpyxl fehlt")
        return
    if not os.path.exists(XLSX):
        lint.fehler.append(f"Knoten-Abgleich: {XLSX} fehlt")
        return
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb["Klimawirkungsketten"]
    kopf = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    soll: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or str(row[0]).strip() != knoten_id:
            continue
        for spalte, wert in zip(kopf, row):
            if spalte and str(spalte).startswith("Input_IDs") and wert:
                soll.update(t.strip() for t in str(wert).split(";") if t.strip())
    wb.close()
    lint.pruefe(bool(soll), f"Knoten-Abgleich {knoten_id}: Inputs gefunden")
    bilanz = src.split("### Knoten-Bilanz")[1].split("### Weitergaben")[0] \
        if "### Knoten-Bilanz" in src else ""
    genannt = set(re.findall(r"\b([ERSW]\d{2,3})\b", bilanz))
    for k in sorted(soll):
        lint.pruefe(k in genannt, f"Knoten-Bilanz enthält {k}",
                    "Knoten der Arbeitsmappe fehlt in der Bilanz")
    for k in sorted(genannt - soll - {knoten_id}):
        lint.pruefe(False, f"Knoten-Bilanz führt {k}",
                    "steht nicht in den Input-Spalten der Arbeitsmappe")
    # KANTEN-Haelfte des §7-Auftrags (Befund 298i): Behauptet der Bericht
    # Output-Kanten, muss die Netzwerkliste sie fuehren — und umgekehrt.
    wb2 = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws2 = wb2["Schadensbaum-Netzwerkliste"]
    kopf2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
    kanten: set[str] = set()
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or str(row[0]).strip() != nr:
            continue
        for spalte, wert in zip(kopf2, row):
            if spalte and ("Output_IDs" in str(spalte)
                           or "Ergänzte Kanten" in str(spalte)) and wert:
                kanten.update(t.strip() for t in str(wert).split(";") if t.strip())
    wb2.close()
    weitergaben = (src.split("### Weitergaben")[1].split("\n## ")[0]
                   if "### Weitergaben" in src else "")
    behauptet_keine = bool(re.search(r"\*\*keine\*\*", weitergaben[:400]))
    # Befund 344(6): Bis Rev. 13 war der Check einseitig — `or not kanten` machte
    # ihn immer gruen, sobald die Mappe keine Kanten fuehrt. Damit fiel eine im
    # BERICHT behauptete Kante, die die Mappe nicht kennt, nie auf. Jetzt beide
    # Richtungen:
    behauptet_kanten = set(re.findall(r"\b([ERSW]\d{2,3})\b", weitergaben))
    lint.pruefe(not (kanten and behauptet_keine), f"Kanten-Abgleich {nr} (Mappe → Bericht)",
                f"Netzwerkliste führt {sorted(kanten)}, Bericht behauptet keine")
    ohne_deckung = sorted(behauptet_kanten - kanten - {knoten_id})
    lint.pruefe(not ohne_deckung, f"Kanten-Abgleich {nr} (Bericht → Mappe)",
                f"Bericht behauptet Kanten {ohne_deckung}, die Netzwerkliste nicht führt")


def revisionshistorie(src: str, lint: Lint) -> None:
    """Jede Revisionsnotiz nennt genau einen — und einen eigenen — Wert (Befund 309/310).

    Die Klasse, die diesen Lauf dominiert hat: Eine globale Ersetzung zieht auch die
    HISTORISCHEN Notizen mit, sodass mehrere Revisionen denselben Wert bei
    verschiedenen Ergebnissen ausweisen. Maschinell pruefbar: Die in der
    Korrekturhistorie genannten Werte muessen paarweise verschieden sein.
    """
    m = re.search(r"\*\*Korrekturhistorie\.\*\*(.*?)(?:\n\n|\Z)", src, re.S)
    if not m:
        lint.ok.append("Revisionshistorie (keine Korrekturhistorie gefunden)")
        return
    paare = re.findall(r"Rev\. (\d+): \**([0-9]+,[0-9]+)\**", m.group(1))
    lint.pruefe(len(paare) >= 2, "Revisionshistorie hat Eintraege")
    werte = [w for _, w in paare]
    doppelt = {w for w in werte if werte.count(w) > 1}
    lint.pruefe(not doppelt, "Revisionshistorie ohne Dubletten",
                f"mehrere Revisionen mit demselben Wert: {sorted(doppelt)} — "
                "typischer Nebeneffekt einer globalen Ersetzung")
    revs = [int(r) for r, _ in paare]
    lint.pruefe(revs == sorted(revs), "Revisionshistorie aufsteigend",
                f"Reihenfolge {revs}")


def pruefe_bericht(pfad: str) -> bool:
    nr = os.path.basename(pfad).split("_")[0]
    src = open(pfad, encoding="utf-8").read()
    lint = Lint()
    UNTERDRUECKT.clear()
    MARKER_ZAEHLER.clear()
    beispiel_bloecke(src, lint)
    zeichentabelle(src, lint)
    werte, baender = parameter_bloecke(src, lint)
    registry_abgleich(nr, werte, lint)
    revisionsrueckstaende(nr, src, baender, lint)
    knoten_abgleich(nr, src, lint)
    revisionshistorie(src, lint)
    abgeloeste_werte(nr, src, lint)
    # Anlagen-Ausgaben ...
    for anlage in ("k_uv_herleitung.md", "ssd_povw.md", "kid2025_baseline.md"):
        pfad_a = os.path.join(ROOT, "data", "kalibrierung", anlage)
        if os.path.exists(pfad_a):
            abgeloeste_werte(nr, open(pfad_a, encoding="utf-8").read(), lint,
                             quelle=f"Anlage {anlage}")
    # ... UND die Quelldateien (Befund 321): Registry, Schadensfunktion,
    # Kalibrierskripte und Golden-Tests tragen dieselben Werte in Kommentaren und
    # source_detail. Bis Rev. 12 las der Lint keine einzige .py-Datei — dort standen
    # deshalb Rueckstaende, die in seiner Negativliste laengst gefuehrt waren.
    for rel in ("app/services/engine/impact/params.py",
                "app/services/engine/impact/health.py",
                "tests/test_methodik_98_golden.py",
                "scripts/kalibrierung/k_uv_herleitung.py",
                "scripts/kalibrierung/kid2025_baseline.py",
                "scripts/kalibrierung/ssd_povw.py",
                # Der Lint prueft sich selbst (Befund 347): Seine Kommentare und
                # Whitelists trugen abgeloeste Werte, waehrend er dieselben Werte
                # im Bericht verbot.
                "scripts/lint_methodik.py"):
        pfad_q = os.path.join(ROOT, rel)
        if os.path.exists(pfad_q):
            abgeloeste_werte(nr, open(pfad_q, encoding="utf-8").read(), lint,
                             quelle=f"Code {os.path.basename(rel)}")
    # Kalibrierskripte muessen wenigstens syntaktisch lauffaehig sein. Anlass: Ein
    # verrutschter Patch hatte kid2025_baseline.py unausfuehrbar gemacht, und das
    # fiel nicht auf — der Lint prueft die ERZEUGTEN Anlagen, nicht die Erzeuger,
    # und die Anlage lag noch vom letzten funktionierenden Lauf vor.
    import ast as _ast
    for rel in ("scripts/kalibrierung/k_uv_herleitung.py",
                "scripts/kalibrierung/kid2025_baseline.py",
                "scripts/kalibrierung/ssd_povw.py"):
        pfad_s = os.path.join(ROOT, rel)
        if not os.path.exists(pfad_s):
            continue
        try:
            _ast.parse(open(pfad_s, encoding="utf-8").read())
            lint.ok.append(f"Kalibrierskript parsebar: {os.path.basename(rel)}")
        except SyntaxError as exc:
            lint.fehler.append(
                f"Kalibrierskript {os.path.basename(rel)} nicht ausfuehrbar: "
                f"Zeile {exc.lineno} — {exc.msg}")

    # Erst wenn alle Quellen gelesen sind, laesst sich sagen, ob eine Ausnahme im
    # gesamten Pruefgut nur eine Stelle deckt (Befund 343).
    quellen_ratchet(lint)

    # Marker-Ratchet (Befund 419): Zahl der markierten Zeilen im Pruefgut muss der
    # bewusst gepflegten Konstante entsprechen — nach oben UND nach unten, damit
    # weder eine stille neue Freigabe noch ein verlorener Marker durchgeht.
    soll = HISTORIE_MARKER_SOLL.get(nr)
    ist = MARKER_ZAEHLER.get(nr, 0)
    if soll is not None:
        lint.pruefe(ist == soll, f"Historie-Marker-Ratchet ({ist} markierte Zeilen)",
                    f"{ist} markierte Zeilen im Prüfgut, HISTORIE_MARKER_SOLL sagt "
                    f"{soll} — bewusst nachtragen, nicht still")

    print(f"\n=== #{nr} · {os.path.basename(pfad)} ===")
    print(f"  {len(lint.ok)} Checks grün")
    gedeckt = UNTERDRUECKT.get(HISTORIE_MARKER, [])
    print(f"  Historie-Marker: {ist} markierte Zeilen (Ratchet {soll}), "
          f"{len(gedeckt)} gedeckte Fundstellen abgelöster Werte:")
    for quelle, zeile, wert in gedeckt:
        print(f"    {quelle} Z. {zeile}: {wert}")
    for f in lint.fehler:
        print(f"  ROT  {f}")
    return not lint.fehler


def main() -> int:
    ziel = sys.argv[1] if len(sys.argv) > 1 else None
    muster = f"{ziel}_*.md" if ziel else "*.md"
    berichte = [p for p in sorted(glob.glob(os.path.join(DOCS, muster)))
                if not p.endswith(".pdf")]
    if not berichte:
        print(f"Kein Bericht gefunden: {os.path.join(DOCS, muster)}")
        return 1
    alle_gruen = all(pruefe_bericht(p) for p in berichte)
    print("\n" + ("ALLE LINTS GRÜN" if alle_gruen else "LINTS ROT"))
    return 0 if alle_gruen else 1


if __name__ == "__main__":
    sys.exit(main())
