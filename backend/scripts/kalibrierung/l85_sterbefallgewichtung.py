#!/usr/bin/env python3
"""L̄_85+ mit Sterbefallgewichten (Methodik #95 Rev. 8; löst Befund 22 / Vermerk I-3).

Die Rev.-5/7-Kette mittelte die Restlebenserwartung 85+ mit BEVÖLKERUNGS-Gewichten
über die Stützstellen e(85)/e(90)/e(95) und kombinierte m/w mit der Bevölkerung —
verlorene Lebensjahre je Sterbefall verlangen aber STERBEFALL-Gewichte (die
Sterbefälle liegen weiter oben in der Altersverteilung). Dieses Skript rechnet
EXAKT mit den realen Sterbefällen nach Einzelaltersjahren:

    L̄_85+,g = [ Σ_{x=85..94} D_x,g · e_g(x)  +  D_95+,g · ē_g(95+) ] / D_85+,g
    ē_g(95+) = Σ_{x≥95} d_x^Tafel · e_g(x) / Σ_{x≥95} d_x^Tafel   (tafelintern
               sterbefallgewichtet — einzige verbleibende, gekennzeichnete
               Näherung: reale 95+-Fälle liegen nur als eine Zeile vor)
    L̄_85+   = Σ_g D_85+,g · L̄_85+,g / Σ_g D_85+,g   (g: männlich/weiblich)

Sensitivität (Bandobergrenze): 95+ mit Untergrenzen-Stützstelle e(95).
Kreuzcheck: Σ Einzeljahre+95+ == Gruppensummen der Tab. 12613-03 (exakt).

Quellen (keyless, Statistische Berichte Destatis):
- Sterbefälle 2023, Tab. 12613-02 (Gestorbene nach Alter, Einzeljahre, m/w;
  Kreuzcheck Tab. 12613-03) — Quelle [49] des Berichts.
- Sterbetafeln 2022/2024, Blätter 12613-b01 (m) / 12613-b02 (w): e(x) und
  Tafel-Gestorbene d_x — Quelle [48] des Berichts.

Ausgaben (Anlagen des Berichts): backend/data/kalibrierung/
    l85_sterbefallgewichtung.csv   (Gruppen, Sterbefälle, e-Stützstellen, Beiträge)
    l85_sterbefallgewichtung.md    (Ergebnis + Gegenrechnung Rev.-7-Werte)

Downloads werden außerhalb des Repos gecacht (~/.cache/kap2-kalibrierung/).
"""
from __future__ import annotations

import csv
import os
import sys
import urllib.request

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DATA = os.path.join(ROOT, "data", "kalibrierung")
CACHE = os.path.expanduser("~/.cache/kap2-kalibrierung")

BASE = ("https://www.destatis.de/DE/Themen/Gesellschaft-Umwelt/Bevoelkerung/"
        "Sterbefaelle-Lebenserwartung/Publikationen/Downloads-Sterbefaelle/")
FILES = {
    "sterbefaelle_2023.xlsx":
        BASE + "statistischer-bericht-sterbefaelle-5126105237005.xlsx"
               "?__blob=publicationFile&v=2",
    "sterbetafeln_2224.xlsx":
        BASE + "statistischer-bericht-sterbetafeln-5126207247005.xlsx"
               "?__blob=publicationFile&v=2",
}

# Gruppen der Tab. 12613-03 im 85+-Bereich → Stützstelle e(x) an der
# Gruppenuntergrenze (identische Stützstellenwahl wie Rev. 5/7, nur um e(100)
# für die eigene 100+-Zeile erweitert).
GROUPS = [("85 - 90", 85), ("90 - 95", 90), ("95 - 100", 95), ("100 u. mehr", 100)]


def _fetch(name: str) -> str:
    os.makedirs(CACHE, exist_ok=True)
    path = os.path.join(CACHE, name)
    if not (os.path.exists(path) and os.path.getsize(path) > 100_000):
        print(f"lade {name} …", file=sys.stderr)
        req = urllib.request.Request(FILES[name], headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=90) as r, open(path, "wb") as fh:
            fh.write(r.read())
    return path


def deaths_85plus_de() -> dict[str, dict[str, float]]:
    """Sterbefälle 2023 Deutschland je 85+-Gruppe und Geschlecht."""
    wb = openpyxl.load_workbook(_fetch("sterbefaelle_2023.xlsx"),
                                read_only=True, data_only=True)
    rows = list(wb["12613-03"].iter_rows(values_only=True))
    hdr = rows[2]
    de_col = next(i for i, c in enumerate(hdr) if str(c).strip() == "Deutschland")
    out: dict[str, dict[str, float]] = {"m": {}, "w": {}}
    sex = None
    for row in rows[3:]:
        label = str(row[0]).strip() if row[0] is not None else ""
        if label == "Männlich":
            sex = "m"
        elif label == "Weiblich":
            sex = "w"
        elif label == "Insgesamt":
            sex = None
        elif sex:
            for grp, _x in GROUPS:
                if label.replace("  ", " ") == grp:
                    out[sex][grp] = float(row[de_col])
    for g in ("m", "w"):
        missing = [grp for grp, _ in GROUPS if grp not in out[g]]
        if missing:
            raise SystemExit(f"Gruppen fehlen ({g}): {missing}")
    return out


def deaths_single_years_de() -> dict[str, tuple[dict[int, float], float]]:
    """Sterbefälle 2023 DE je Einzelaltersjahr 85–94 + 95+-Rest, je Geschlecht.

    Tab. 12613-02 (Spalten: 1 = Insgesamt, 4 = Männlich, 7 = Weiblich); letzte
    Alterszeile „95 und älter einschl. unbekannt"."""
    import re as _re
    wb = openpyxl.load_workbook(_fetch("sterbefaelle_2023.xlsx"),
                                read_only=True, data_only=True)
    rows = list(wb["12613-02"].iter_rows(values_only=True))
    out: dict[str, tuple[dict[int, float], float]] = {}
    for g, col in (("m", 4), ("w", 7)):
        single: dict[int, float] = {}
        rest = None
        for row in rows:
            label = str(row[0]).strip() if row[0] is not None else ""
            m = _re.match(r"^(\d+)\s*-\s*\d+$", label)
            if m and int(m.group(1)) >= 85:
                single[int(m.group(1))] = float(row[col])
            elif label.startswith("95 und älter"):
                rest = float(row[col])
        if sorted(single) != list(range(85, 95)) or rest is None:
            raise SystemExit(f"Einzeljahre 85–94/95+ unvollständig ({g}): {sorted(single)}")
        out[g] = (single, rest)
    return out


def life_expectancy() -> dict[str, dict[int, float]]:
    """e(x) der Sterbetafel 2022/2024 je Geschlecht und Altersjahr."""
    wb = openpyxl.load_workbook(_fetch("sterbetafeln_2224.xlsx"),
                                read_only=True, data_only=True)
    out: dict[str, dict[int, float]] = {}
    for g, sheet in (("m", "12613-b01"), ("w", "12613-b02")):
        ex: dict[int, float] = {}
        dx: dict[int, float] = {}
        for row in wb[sheet].iter_rows(values_only=True):
            try:
                x = int(row[0])
            except (TypeError, ValueError):
                continue
            ex[x] = float(row[7])
            dx[x] = float(row[4])
        out[g] = ex
        out[g + "_dx"] = dx
    return out


def main() -> None:
    groups = deaths_85plus_de()          # Kreuzcheck (Tab. 12613-03)
    single = deaths_single_years_de()    # Basis (Tab. 12613-02)
    ex = life_expectancy()

    rows_csv: list[dict] = []
    l_by_sex: dict[str, float] = {}
    l_by_sex_sens: dict[str, float] = {}
    d_by_sex: dict[str, float] = {}
    e95p: dict[str, float] = {}
    for g in ("m", "w"):
        years, rest95 = single[g]
        # Kreuzcheck gegen die Gruppentabelle — muss exakt aufgehen.
        d_total = sum(years.values()) + rest95
        d_groups = sum(groups[g][grp] for grp, _ in GROUPS)
        if abs(d_total - d_groups) > 0.5:
            raise SystemExit(f"Kreuzcheck 12613-02 vs. -03 verletzt ({g}): "
                             f"{d_total} != {d_groups}")
        # ē(95+): tafelintern sterbefallgewichtet (d_x der Sterbetafel als
        # gruppeninterne Verteilung — gekennzeichnete Restnäherung).
        dx = ex[g + "_dx"]
        xs95 = [x for x in dx if x >= 95]
        e95p[g] = (sum(dx[x] * ex[g][x] for x in xs95)
                   / sum(dx[x] for x in xs95))
        num = sum(d * ex[g][x] for x, d in years.items()) + rest95 * e95p[g]
        num_sens = sum(d * ex[g][x] for x, d in years.items()) + rest95 * ex[g][95]
        l_by_sex[g] = num / d_total
        l_by_sex_sens[g] = num_sens / d_total
        d_by_sex[g] = d_total
        for x, d in sorted(years.items()):
            rows_csv.append({"geschlecht": g, "alter": x,
                             "sterbefaelle_2023_de": round(d),
                             "e_x_2224": round(ex[g][x], 4),
                             "beitrag_d_mal_e": round(d * ex[g][x], 1)})
        rows_csv.append({"geschlecht": g, "alter": "95+",
                         "sterbefaelle_2023_de": round(rest95),
                         "e_x_2224": round(e95p[g], 4),
                         "beitrag_d_mal_e": round(rest95 * e95p[g], 1)})

    d_all = sum(d_by_sex.values())
    l_comb = sum(d_by_sex[g] * l_by_sex[g] for g in ("m", "w")) / d_all
    l_comb_sens = sum(d_by_sex[g] * l_by_sex_sens[g] for g in ("m", "w")) / d_all

    os.makedirs(DATA, exist_ok=True)
    with open(os.path.join(DATA, "l85_sterbefallgewichtung.csv"), "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows_csv[0]))
        w.writeheader()
        w.writerows(rows_csv)

    rev7 = 5.44
    lines = [
        "# L̄_85+ mit Sterbefallgewichten (#95 Rev. 8 — Befund 22 / Vermerk I-3)",
        "",
        "Exakte Rechnung: reale Sterbefälle 2023 nach Einzelaltersjahren 85–94 ×",
        "e(x) der Sterbetafel 2022/2024; 95+-Restzeile mit tafelintern",
        "sterbefallgewichtetem ē(95+) (gekennzeichnete Restnäherung). Kreuzcheck",
        "gegen die Gruppentabelle 12613-03: exakt bestanden. Quellen [48], [49].",
        "",
        "- Sterbefälle 85+ (2023, DE): männlich {} · weiblich {} · gesamt {}".format(
            *(f"{d_by_sex['m']:,.0f} {d_by_sex['w']:,.0f} {d_all:,.0f}"
              .replace(",", ".").split())),
        f"- ē(95+): männlich {e95p['m']:.3f} · weiblich {e95p['w']:.3f} J "
        f"(zum Vergleich Stützstelle e(95): {ex['m'][95]:.2f} / {ex['w'][95]:.2f})",
        f"- **L̄_85+ männlich = {l_by_sex['m']:.3f} J** · "
        f"**weiblich = {l_by_sex['w']:.3f} J**",
        f"- **L̄_85+ kombiniert (sterbefallgewichtet m/w) = {l_comb:.3f} J** "
        f"→ Berichtswert gerundet **{l_comb:.2f}**",
        f"- Sensitivität 95+-Stützstelle e(95) statt ē(95+): {l_comb_sens:.3f} J "
        f"→ Band **[{l_comb:.2f}, {l_comb_sens:.2f}]** (Untergrenze = Basiswert; "
        f"die e(95)-Stützstelle überschätzt, e fällt mit x)",
        f"- Rev.-7-Wert (Bevölkerungsgewichte, Stützstellen): {rev7} J → "
        f"Differenz {l_comb - rev7:+.2f} J (§3.5 hatte −0,3…−0,5 J geschätzt; der "
        f"Mehrbetrag stammt aus der sterbefallgewichteten m/w-Kombination und den "
        f"Einzeljahren oberhalb der Gruppen-Untergrenzen)",
    ]
    with open(os.path.join(DATA, "l85_sterbefallgewichtung.md"), "w") as fh:
        fh.write("\n".join(lines) + "\n")
    print("\n".join(lines))


if __name__ == "__main__":
    main()
