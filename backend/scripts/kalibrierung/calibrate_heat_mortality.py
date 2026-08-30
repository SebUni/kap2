"""Empirische Kalibrierung c_kal der Hitzemortalitäts-ERF + Bundesland-Verteilungsprüfung
(Grundsätze G5/G12/G14, docs/METHODIK_GRUNDSAETZE.md).

Nationaler Näherungslauf auf Bundesland-Ebene (ohne 100-m-Gitter, ohne UHI — der UHI-Zuschlag
ist je 1-km-Zelle mittelwerttreu und hebt die Summe nur über die Konvexität der Exponential-
kurve geringfügig; die Näherung ist damit konservativ):

    D_mod(J, BL) = Σ_a pop_{BL,a} · m_a/1e5 · (1/52) · Σ_w [ exp(β_a·(T_{BL,J} + q_{w,reg} − T0_reg)_+) − 1 ]
    c_kal = Σ_J D_RKI(J)·D_mod(J) / Σ_J D_mod(J)²        (Kleinste Quadrate durch den Ursprung)

Anker: RKI Epid Bull 19/2025, Anhang 1 (CC BY 4.0): Erwartungswerte + Prädiktionsintervalle für
Deutschland und die Bundesländer 1992–2024; signifikant = untere Prädiktionsgrenze > 0.
2025 (≈ 2.500, RKI-Wochenbericht KW 38/2025) wird ergänzt; 2026 (laufende Saison) ausgeschlossen.

Eingaben (backend/data/kalibrierung/):
  rki_eb19_2025_anhang_bundeslaender.xlsx   Sheets „Deutschland", „Bundesländer"
  sommermittel_bundesland.csv               jahr, bundesland, t_sommer        (DWD-Gebietsmittel Jun–Aug)
  wochenquantile_region.csv                 region, w, p, q_w_emp, q_w_gauss  (DWD-Stationen 1991–2020)
  bevoelkerung_bundesland_altersband.csv    bundesland, u65, a65_74, a75_84, a85p (Destatis 12411)
Ausgabe: c_kal_ergebnis_{emp|gauss}.csv/.md, bundesland_validierung_{emp|gauss}.csv
"""
from __future__ import annotations

import csv
import math
import os
import sys
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, ROOT)
DATA = os.path.join(ROOT, "data", "kalibrierung")

from app.services.engine.impact.health import (  # noqa: E402
    AGE_BANDS, AGE_BETA_FACTOR, REGION_BETA_85P, REGION_BY_BUNDESLAND, REGION_THRESHOLD,
)
from app.data.germany_health_reference import BASELINE_MORTALITY_PER_100K as _REPO_M_A  # noqa: E402

# Basissterberaten 2023 je Altersband (Destatis Sterbefälle 2023 Tab. 12613-03 ÷ Bevölkerung 31.12.2023;
# Herleitung in docs/METHODIK_M0_GESUNDHEIT.pdf Kap. 2). Repo-Konstanten (180/1800/4600/15500) werden in AP4
# nachgezogen; u65 weicht dort um +18 % ab.
BASELINE_MORTALITY_PER_100K = {"u65": 213.2, "a65_74": 1737.9, "a75_84": 4812.3, "a85p": 14800.2}

RKI_2025 = (2025, 2_500, 1_200, 3_700)   # RKI-Wochenbericht KW 38/2025 (Saisonabschluss)
VALIDATION_YEARS = (2003, 2018, 2019, 2022)   # deutliche Hitzejahre für die Bundesland-Prüfung

# Vier-Regionen-Zuschnitt der revidierten RKI-Reihe (Epid Bull 19/2025: Norden/Osten/Westen/Süden);
# Kurvenparameter bleiben die der Winklmayr-2022-Region (Osten/Westen = „mitte"), nur der
# Kalibrierfaktor wird je RKI-Region bestimmt.
RKI4_BY_BUNDESLAND = {
    "Schleswig-Holstein": "norden", "Hamburg": "norden", "Mecklenburg-Vorpommern": "norden",
    "Niedersachsen": "norden", "Bremen": "norden",
    "Brandenburg": "osten", "Berlin": "osten", "Sachsen": "osten", "Sachsen-Anhalt": "osten", "Thüringen": "osten",
    "Nordrhein-Westfalen": "westen", "Hessen": "westen", "Rheinland-Pfalz": "westen", "Saarland": "westen",
    "Baden-Württemberg": "sueden", "Bayern": "sueden",
}


def _read(name: str) -> list[dict]:
    with open(os.path.join(DATA, name), encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _rki() -> tuple[dict[int, tuple[int, int, int]], dict[tuple[int, str], tuple[int, int, int]]]:
    wb = openpyxl.load_workbook(os.path.join(DATA, "rki_eb19_2025_anhang_bundeslaender.xlsx"), data_only=True)
    de, bl = {}, {}
    for row in list(wb["Deutschland"].iter_rows(values_only=True))[1:]:
        if isinstance(row[0], (int, float)):
            de[int(row[0])] = (int(row[1]), int(row[2]), int(row[3]))
    for row in list(wb["Bundesländer"].iter_rows(values_only=True))[1:]:
        if isinstance(row[0], (int, float)):
            bl[(int(row[0]), str(row[1]))] = (int(row[2]), int(row[3]), int(row[4]))
    de[RKI_2025[0]] = RKI_2025[1:]
    return de, bl


def model_year(J: int, t_sommer, q_w, pop) -> dict[str, float]:
    out = {}
    for bl_name, bands in pop.items():
        reg = REGION_BY_BUNDESLAND[bl_name]
        t0, beta85 = REGION_THRESHOLD[reg], REGION_BETA_85P[reg]
        t_bar = t_sommer[(J, bl_name)]
        total = 0.0
        for band in AGE_BANDS:
            beta = beta85 * AGE_BETA_FACTOR[band]
            excess = sum(math.exp(beta * max(0.0, t_bar + q - t0)) - 1.0 for q in q_w[reg])
            total += bands[band] * BASELINE_MORTALITY_PER_100K[band] / 1e5 * (1.0 / 52.0) * excess
        out[bl_name] = total
    return out


def main(use_gauss: bool = False, year_from: int = 1992, regions4: bool = False) -> None:
    tag = ("gauss" if use_gauss else "emp") + (f"_ab{year_from}" if year_from > 1992 else "") + ("_4reg" if regions4 else "")
    REG = RKI4_BY_BUNDESLAND if regions4 else REGION_BY_BUNDESLAND
    de, bl_rki = _rki()
    t_sommer = {(int(r["jahr"]), r["bundesland"]): float(r["t_sommer"]) for r in _read("sommermittel_bundesland.csv")}
    qkey = "q_w_gauss" if use_gauss else "q_w_emp"
    q_w: dict[str, list[float]] = defaultdict(list)
    for r in sorted(_read("wochenquantile_region.csv"), key=lambda r: (r["region"], int(r["w"]))):
        q_w[r["region"]].append(float(r[qkey]))
    pop = {r["bundesland"]: {b: float(r[b]) for b in AGE_BANDS} for r in _read("bevoelkerung_bundesland_altersband.csv")}

    years = sorted(J for J, (ew, lo, hi) in de.items() if lo > 0 and J >= year_from and all((J, b) in t_sommer for b in pop))
    per_year = {J: model_year(J, t_sommer, q_w, pop) for J in years}
    d_mod = [sum(per_year[J].values()) for J in years]
    d_rki = [de[J][0] for J in years]
    c_kal = sum(a * b for a, b in zip(d_rki, d_mod)) / sum(b * b for b in d_mod)
    mean_rki = sum(d_rki) / len(d_rki)
    r2 = 1 - sum((a - c_kal * b) ** 2 for a, b in zip(d_rki, d_mod)) / sum((a - mean_rki) ** 2 for a in d_rki)

    rows = []
    for J, m, r in zip(years, d_mod, d_rki):
        rows.append({"jahr": J, "d_rki": r, "pi_unten": de[J][1], "pi_oben": de[J][2],
                     "d_mod_unkalibriert": round(m), "d_mod_kalibriert": round(c_kal * m),
                     "abweichung_pct": round(100 * (c_kal * m - r) / r, 1),
                     "im_praediktionsintervall": de[J][1] <= c_kal * m <= de[J][2]})
    with open(os.path.join(DATA, f"c_kal_ergebnis_{tag}.csv"), "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys())); w.writeheader(); w.writerows(rows)

    # Regionale Kalibrierfaktoren (G12): LS durch Ursprung über alle Bundesland-Jahre im Fenster.
    all_years = sorted(J for J in de if J >= year_from and J <= 2024 and all((J, b) in t_sommer for b in pop))
    per_year_all = {J: per_year.get(J) or model_year(J, t_sommer, q_w, pop) for J in all_years}
    c_reg = {}
    for reg in set(REG[b] for b in pop):
        num = den = 0.0
        for J in all_years:
            for b in pop:
                if REG[b] != reg or (J, b) not in bl_rki:
                    continue
                m = per_year_all[J][b]; num += bl_rki[(J, b)][0] * m; den += m * m
        c_reg[reg] = num / den if den else float('nan')

    # Bundesland-Verteilungsprüfung (G12): Summe über Hitzejahre, Modell × c_kal vs. RKI.
    val = []
    for bl_name in pop:
        m = sum(c_kal * per_year[J][bl_name] for J in VALIDATION_YEARS if J in per_year)
        r = sum(bl_rki[(J, bl_name)][0] for J in VALIDATION_YEARS if (J, bl_name) in bl_rki)
        p = sum(pop[bl_name].values())
        m_reg = sum(c_reg[REG[bl_name]] * per_year[J][bl_name] for J in VALIDATION_YEARS if J in per_year)
        val.append({"bundesland": bl_name, "region": REG[bl_name],
                    "rki_summe": r, "modell_summe": round(m),
                    "rki_je_100k": round(1e5 * r / p / len(VALIDATION_YEARS), 1),
                    "modell_je_100k": round(1e5 * m / p / len(VALIDATION_YEARS), 1),
                    "verhaeltnis_modell_rki": round(m / r, 2) if r else None,
                    "modell_regional_je_100k": round(1e5 * m_reg / p / len(VALIDATION_YEARS), 1),
                    "verhaeltnis_regional": round(m_reg / r, 2) if r else None})
    val.sort(key=lambda x: x["rki_je_100k"])
    with open(os.path.join(DATA, f"bundesland_validierung_{tag}.csv"), "w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(val[0].keys())); w.writeheader(); w.writerows(val)

    n_in = sum(1 for r in rows if r["im_praediktionsintervall"])
    lines = [f"# c_kal — empirische Kalibrierung ({'Gauß-Quantile' if use_gauss else 'empirische Wochenquantile'})", "",
             f"c_kal = **{c_kal:.3f}** (Kleinste Quadrate durch Ursprung, {len(years)} signifikante Jahre) · R² = {r2:.3f} · "
             f"{n_in}/{len(years)} Jahre innerhalb des RKI-Prädiktionsintervalls", "",
             "| Jahr | RKI [PI] | Modell roh | Modell × c_kal | Abw. % | im PI |", "|---|---|---|---|---|---|"]
    lines += [f"| {r['jahr']} | {r['d_rki']} [{r['pi_unten']}–{r['pi_oben']}] | {r['d_mod_unkalibriert']} | {r['d_mod_kalibriert']} | {r['abweichung_pct']:+.1f} | {'ja' if r['im_praediktionsintervall'] else 'nein'} |" for r in rows]
    lines += ["", f"## Bundesland-Verteilungsprüfung (Σ {', '.join(map(str, VALIDATION_YEARS))}, je 100.000 EW und Jahr)", "",
              "| Bundesland | Region | RKI | Modell (c_kal) | Verh. | Modell (c_reg) | Verh. |", "|---|---|---|---|---|---|---|"]
    lines += [f"| {v['bundesland']} | {v['region']} | {v['rki_je_100k']} | {v['modell_je_100k']} | {v['verhaeltnis_modell_rki']} | {v['modell_regional_je_100k']} | {v['verhaeltnis_regional']} |" for v in val]
    lines += ["", "Regionale Kalibrierfaktoren c_reg (LS über Bundesland-Jahre " + f"{all_years[0]}–{all_years[-1]}): " + ", ".join(f"{k} = {v:.3f}" for k, v in sorted(c_reg.items()))]
    with open(os.path.join(DATA, f"c_kal_ergebnis_{tag}.md"), "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")
    print("\n".join(lines))


if __name__ == "__main__":
    yf = 1992
    for a in sys.argv[1:]:
        if a.startswith("--from="):
            yf = int(a.split("=")[1])
    main(use_gauss="--gauss" in sys.argv, year_from=yf, regions4="--regions4" in sys.argv)
